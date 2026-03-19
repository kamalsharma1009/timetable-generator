from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, time
import json
import os
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import random
import copy
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or 'dev-key-12345'
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL') or 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Timetable settings
app.config['DAYS'] = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
app.config['TIME_SLOTS'] = [
            ('9:40', '10:40'),
            ('10:40', '11:40'),
            ('11:40', '12:40'),
            ('12:40', '1:10'),  # Lunch Break
            ('1:10', '2:10'),
            ('2:10', '3:10'),
            ('3:10', '3:20'),  # Short Break
            ('3:20', '4:20'),
            ('4:20', '5:20')
]
app.config['LECTURE_SLOT_INDICES'] = [0, 1, 3, 4, 5, 6]
app.config['PRACTICAL_SLOTS'] = [('14:00', '16:00'), ('15:00', '17:10')]

# Genetic Algorithm parameters
app.config['POPULATION_SIZE'] = 100
app.config['GENERATIONS'] = 500
app.config['MUTATION_RATE'] = 0.1
app.config['ELITE_SIZE'] = 20

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Models (defined here to avoid circular imports)
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200))
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Department(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(10), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    classes = db.relationship('Class', backref='department', lazy=True)
    faculty = db.relationship('Faculty', backref='department', lazy=True)

class Class(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    code = db.Column(db.String(20), unique=True, nullable=False)
    year = db.Column(db.String(10), nullable=False)  # FY, SY, TY
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'), nullable=False)
    semester = db.Column(db.Integer, nullable=False)
    strength = db.Column(db.Integer, default=60)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    subjects = db.relationship('ClassSubject', backref='class_ref', lazy=True)
    batches = db.relationship('Batch', backref='class_ref', lazy=True)
    timetable = db.relationship('Timetable', backref='class_ref', lazy=True)

class Batch(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(10), nullable=False)  # TB1, TB2, TB3
    code = db.Column(db.String(30), unique=True, nullable=False)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    mentor_id = db.Column(db.Integer, db.ForeignKey('faculty.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    practical_slots = db.relationship('PracticalSlot', backref='batch', lazy=True)

class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(200), nullable=False)
    type = db.Column(db.String(20), nullable=False)  # Theory, Practical, Lab, Tutorial
    lecture_hours = db.Column(db.Integer, default=0)
    practical_hours = db.Column(db.Integer, default=0)
    credits = db.Column(db.Integer, default=3)
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    class_subjects = db.relationship('ClassSubject', backref='subject_ref', lazy=True)

class ClassSubject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.id'))
    lecture_slots_per_week = db.Column(db.Integer, default=3)
    practical_slots_per_week = db.Column(db.Integer, default=2)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Faculty(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    phone = db.Column(db.String(15))
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'))
    designation = db.Column(db.String(50))
    qualification = db.Column(db.String(100))
    max_hours_per_day = db.Column(db.Integer, default=4)
    max_hours_per_week = db.Column(db.Integer, default=20)
    availability = db.Column(db.Text, default='{}')  # JSON storing availability per day
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    class_subjects = db.relationship('ClassSubject', backref='faculty_ref', lazy=True)
    mentored_batches = db.relationship('Batch', backref='mentor', lazy=True)
    timetable_entries = db.relationship('Timetable', backref='faculty_ref', lazy=True)

class Room(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    room_number = db.Column(db.String(20), unique=True, nullable=False)
    room_type = db.Column(db.String(20), nullable=False)  # Classroom, Lab, Auditorium
    capacity = db.Column(db.Integer, nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'))
    equipment = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    timetable_entries = db.relationship('Timetable', backref='room', lazy=True)

class Timetable(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    day = db.Column(db.String(10), nullable=False)
    slot_number = db.Column(db.Integer, nullable=False)
    start_time = db.Column(db.String(10), nullable=False)
    end_time = db.Column(db.String(10), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'))
    subject = db.relationship('Subject')
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.id'))
    # faculty_ref backref is set by Faculty.timetable_entries (access via entry.faculty_ref)
    room_id = db.Column(db.Integer, db.ForeignKey('room.id'))
    # room backref is set by Room.timetable_entries (access via entry.room)
    batch_id = db.Column(db.Integer, db.ForeignKey('batch.id'))
    batch = db.relationship('Batch')
    session_type = db.Column(db.String(20), nullable=False)  # Lecture, Practical, Mentoring, Break
    is_break = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class PracticalSlot(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    batch_id = db.Column(db.Integer, db.ForeignKey('batch.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.id'))
    room_id = db.Column(db.Integer, db.ForeignKey('room.id'))
    day = db.Column(db.String(10), nullable=False)
    start_time = db.Column(db.String(10), nullable=False)
    end_time = db.Column(db.String(10), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# Create tables before first request
with app.app_context():
    db.create_all()
    # Create default admin user if not exists
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', email='admin@college.edu', is_admin=True)
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print("Default admin user created: username='admin', password='admin123'")

# Genetic Algorithm Class (moved here to avoid circular imports)
class GeneticAlgorithmTimetable:
    def __init__(self, class_id):
        self.class_id = class_id
        self.class_obj = Class.query.get(class_id)
        self.days = app.config['DAYS']
        self.time_slots = [
            ('09:40', '10:40'),
            ('10:40', '11:40'),
            ('11:40', '12:40'),
            ('12:40', '13:10'),  # Lunch Break  (index 3)
            ('13:10', '14:10'),
            ('14:10', '15:10'),
            ('15:10', '15:20'),  # Short Break  (index 6)
            ('15:20', '16:20'),
            ('16:20', '17:20')
        ]
        app.config['TIME_SLOTS'] = self.time_slots
        # Break slot indices
        self.BREAK_SLOTS = {3, 6}
        # All valid (non-break) slot indices for lectures and practicals
        self.valid_slots = [i for i in range(len(self.time_slots)) if i not in self.BREAK_SLOTS]
        # Consecutive slot pairs allowed for 2-slot practicals (both slots must be non-break and consecutive)
        self.practical_pairs = [
            (s, s+1) for s in self.valid_slots
            if (s+1) in self.valid_slots and (s+1 - s == 1)
        ]
        self.lecture_slots = app.config['LECTURE_SLOT_INDICES']

        # Get all required data
        self.class_subjects = ClassSubject.query.filter_by(class_id=class_id).all()
        self.batches = Batch.query.filter_by(class_id=class_id).all()
        self.faculty_list = Faculty.query.all()
        dept_id = self.class_obj.department_id

        self.rooms = Room.query.filter_by(department_id=dept_id).all()
        self.classrooms = [r for r in self.rooms if r.room_type == 'Classroom']
        self.labs = [r for r in self.rooms if r.room_type == 'Lab']

        if not self.classrooms:
            raise Exception("No classrooms available for this department")
        if not self.labs:
            raise Exception("No labs available for this department")

        # Initialize population
        self.population = []
        self.best_solution = None
        self.best_fitness = float('-inf')
        
    def create_individual(self):
        """Create one feasible timetable individual"""
        individual = {
            'lectures': [],
            'practicals': [],
            'mentoring': []
        }
        
        # Try to create a feasible individual
        max_attempts = 50
        for attempt in range(max_attempts):
            try:
                # 1. Schedule lectures
                for class_subject in self.class_subjects:
                    lecture_count = class_subject.lecture_slots_per_week
                    scheduled = 0
                    
                    attempts = 0
                    max_attempts_per_subject = 50

                    while scheduled < lecture_count and attempts < max_attempts_per_subject:
                        attempts += 1
                        day = random.choice(self.days)
                        slot_idx = random.choice(self.lecture_slots)
                        
                        # Skip lunch break
                        if slot_idx == 2:
                            continue
                            
                        # Check faculty availability
                        faculty_busy = self.check_faculty_busy(individual, class_subject.faculty_id, day, slot_idx)
                        if faculty_busy:
                            continue
                        
                        # Find available classroom
                        classroom = self.find_available_classroom(individual, day, slot_idx)
                        if not classroom:
                            continue
                        
                        slot = self.time_slots[slot_idx]
                        individual['lectures'].append({
                            'class_subject_id': class_subject.id,
                            'subject_id': class_subject.subject_id,
                            'faculty_id': class_subject.faculty_id,
                            'room_id': classroom.id,
                            'day': day,
                            'slot_number': slot_idx,
                            'start_time': slot[1],
                            'end_time': slot[2],
                            'session_type': 'Lecture'
                        })
                        scheduled += 1
                    if scheduled < lecture_count:
                        print("Warning: Could not schedule all lectures")
                # 2. Schedule practicals for each batch
                for batch in self.batches:
                    for class_subject in self.class_subjects:
                        practical_count = class_subject.practical_slots_per_week
                        scheduled = 0
                        
                        attempts = 0
                        max_attempts_per_subject = 50

                        while scheduled < practical_count and attempts < max_attempts_per_subject:
                            attempts += 1
                            # Schedule 2-hour practical in afternoon
                            day = random.choice(self.days)
                            # Use slots 4-5 or 5-6 for practicals (afternoon)
                            start_slot = random.choice([4, 5])
                            end_slot = start_slot + 1
                            
                            if end_slot > 6:  # Ensure within bounds
                                continue
                            
                            # Check faculty availability for both slots
                            faculty_busy = False
                            for slot in [start_slot, end_slot]:
                                if self.check_faculty_busy(individual, class_subject.faculty_id, day, slot):
                                    faculty_busy = True
                                    break
                            
                            if faculty_busy:
                                continue
                            
                            # Find available lab
                            lab = self.find_available_lab(individual, day, start_slot, end_slot)
                            if not lab:
                                continue
                            
                            individual['practicals'].append({
                                'batch_id': batch.id,
                                'subject_id': class_subject.subject_id,
                                'faculty_id': class_subject.faculty_id,
                                'room_id': lab.id,
                                'day': day,
                                'start_slot': start_slot,
                                'end_slot': end_slot,
                                'start_time': self.time_slots[start_slot][1],
                                'end_time': self.time_slots[end_slot][2],
                                'session_type': 'Practical'
                            })
                            scheduled += 1
                        if scheduled < practical_count:
                            print("Warning: Could not schedule all practicals")
                # 3. Schedule mentoring sessions
                for batch in self.batches:
                    if batch.mentor_id:
                        for _ in range(20):  # Try 20 times
                            day = random.choice(self.days)
                            slot_idx = random.choice(self.lecture_slots)
                            
                            # Skip lunch break
                            if slot_idx == 2:
                                continue
                                
                            # Check mentor availability
                            if self.check_faculty_busy(individual, batch.mentor_id, day, slot_idx):
                                continue
                            
                            classroom = self.find_available_classroom(individual, day, slot_idx)
                            if not classroom:
                                continue
                            
                            slot = self.time_slots[slot_idx]
                            individual['mentoring'].append({
                                'batch_id': batch.id,
                                'faculty_id': batch.mentor_id,
                                'room_id': classroom.id,
                                'day': day,
                                'slot_number': slot_idx,
                                'start_time': slot[1],
                                'end_time': slot[2],
                                'session_type': 'Mentoring'
                            })
                            break
                
                return individual
                
            except Exception as e:
                # Reset and try again
                individual = {'lectures': [], 'practicals': [], 'mentoring': []}
                continue
        
        # If we couldn't create a feasible individual after max attempts, return what we have
        return individual
    
    def check_faculty_busy(self, individual, faculty_id, day, slot_number):
        """Check if faculty is already busy at given day and slot"""
        if not faculty_id:
            return False
            
        # Check lectures
        for lecture in individual['lectures']:
            if (lecture['faculty_id'] == faculty_id and 
                lecture['day'] == day and 
                lecture['slot_number'] == slot_number):
                return True
        
        # Check practicals
        for practical in individual['practicals']:
            if practical['faculty_id'] == faculty_id and practical['day'] == day:
                if slot_number >= practical['start_slot'] and slot_number <= practical['end_slot']:
                    return True
        
        # Check mentoring
        for mentoring in individual['mentoring']:
            if (mentoring['faculty_id'] == faculty_id and 
                mentoring['day'] == day and 
                mentoring['slot_number'] == slot_number):
                return True
        
        return False
    
    def find_available_classroom(self, individual, day, slot_number):
        """Find available classroom for given time slot"""
        if not self.classrooms:
            return None
            
        available_classrooms = self.classrooms.copy()
        
        # Remove classrooms that are occupied
        for lecture in individual['lectures']:
            if lecture['day'] == day and lecture['slot_number'] == slot_number:
                available_classrooms = [r for r in available_classrooms if r.id != lecture['room_id']]
        
        for mentoring in individual['mentoring']:
            if mentoring['day'] == day and mentoring['slot_number'] == slot_number:
                available_classrooms = [r for r in available_classrooms if r.id != mentoring['room_id']]
        
        return random.choice(available_classrooms) if available_classrooms else None
    
    def find_available_lab(self, individual, day, start_slot, end_slot):
        """Find available lab for 2-hour practical"""
        if not self.labs:
            return None
            
        available_labs = self.labs.copy()
        
        # Remove labs that are occupied during these slots
        for practical in individual['practicals']:
            if practical['day'] == day:
                # Check for overlap
                if not (end_slot < practical['start_slot'] or start_slot > practical['end_slot']):
                    available_labs = [r for r in available_labs if r.id != practical['room_id']]
        
        return random.choice(available_labs) if available_labs else None
    
    def calculate_fitness(self, individual):
        """Calculate fitness score for individual"""
        fitness = 1000  # Base score
        
        # Penalties for violations
        penalties = 0
        
        # 1. Check faculty overload
        faculty_hours = {}
        for lecture in individual['lectures']:
            fid = lecture['faculty_id']
            if fid:
                faculty_hours[fid] = faculty_hours.get(fid, 0) + 1
        
        for practical in individual['practicals']:
            fid = practical['faculty_id']
            if fid:
                faculty_hours[fid] = faculty_hours.get(fid, 0) + 2
        
        for mentoring in individual['mentoring']:
            fid = mentoring['faculty_id']
            if fid:
                faculty_hours[fid] = faculty_hours.get(fid, 0) + 1
        
        # Check against faculty max hours (simplified)
        for fid, hours in faculty_hours.items():
            if hours > 20:  # Default max hours
                penalties += (hours - 20) * 10
        
        # 2. Check room conflicts
        room_schedule = {}
        for lecture in individual['lectures']:
            key = (lecture['room_id'], lecture['day'], lecture['slot_number'])
            if key in room_schedule:
                penalties += 50
            room_schedule[key] = True
        
        # 3. Check if all subjects have required hours (simplified)
        subject_lecture_hours = {}
        subject_practical_hours = {}
        
        for lecture in individual['lectures']:
            sid = lecture['subject_id']
            subject_lecture_hours[sid] = subject_lecture_hours.get(sid, 0) + 1
        
        for practical in individual['practicals']:
            sid = practical['subject_id']
            subject_practical_hours[sid] = subject_practical_hours.get(sid, 0) + 2
        
        for class_subject in self.class_subjects:
            required_lecture = class_subject.lecture_slots_per_week
            required_practical = class_subject.practical_slots_per_week * 2  # 2 hours each
            
            actual_lecture = subject_lecture_hours.get(class_subject.subject_id, 0)
            actual_practical = subject_practical_hours.get(class_subject.subject_id, 0)
            
            if actual_lecture < required_lecture:
                penalties += (required_lecture - actual_lecture) * 5
            if actual_practical < required_practical:
                penalties += (required_practical - actual_practical) * 5
        
        # 4. Reward for having all batches with mentoring
        mentoring_count = len(individual['mentoring'])
        required_mentoring = len([b for b in self.batches if b.mentor_id])
        if mentoring_count < required_mentoring:
            penalties += (required_mentoring - mentoring_count) * 100
        
        fitness -= penalties
        return max(fitness, 0)  # Ensure fitness is not negative
    
    def crossover(self, parent1, parent2):
        """Create child through crossover"""
        child = {
            'lectures': [],
            'practicals': [],
            'mentoring': []
        }
        
        # Simple crossover: take half from each parent
        if parent1['lectures'] and parent2['lectures']:
            split = len(parent1['lectures']) // 2
            child['lectures'] = parent1['lectures'][:split] + parent2['lectures'][split:]
        
        if parent1['practicals'] and parent2['practicals']:
            split = len(parent1['practicals']) // 2
            child['practicals'] = parent1['practicals'][:split] + parent2['practicals'][split:]
        
        if parent1['mentoring'] and parent2['mentoring']:
            split = len(parent1['mentoring']) // 2
            child['mentoring'] = parent1['mentoring'][:split] + parent2['mentoring'][split:]
        
        return child
    
    def mutate(self, individual):
        """Apply mutation to individual"""
        if individual['lectures'] and random.random() < 0.3:
            # Mutate a random lecture
            idx = random.randint(0, len(individual['lectures']) - 1)
            lecture = individual['lectures'][idx]
            
            # Change day or slot
            if random.random() < 0.5:
                lecture['day'] = random.choice(self.days)
            else:
                new_slot = random.choice([s for s in self.lecture_slots if s != 2])  # Skip lunch
                lecture['slot_number'] = new_slot
                slot = self.time_slots[new_slot]
                lecture['start_time'] = slot[1]
                lecture['end_time'] = slot[2]
        
        return individual
    
    def generate_structured_timetable(self):
        """
        Constraint-compliant timetable generator.
        Order: Breaks → Practicals (parallel) → Mentoring → Lectures
        """
        # ------------------------------------------------------------------
        # RESET
        # ------------------------------------------------------------------
        Timetable.query.filter_by(class_id=self.class_id).delete()
        db.session.commit()

        days = self.days
        slots = self.time_slots  # list of (start, end) tuples, len=9
        n_batches = len(self.batches)

        # Busy trackers  key → True
        faculty_busy = {}   # (faculty_id, day, slot_idx)
        room_busy    = {}   # (room_id,    day, slot_idx)
        class_busy   = {}   # (day, slot_idx)  — at most one activity per slot per class
        
        # Pre-load busy faculty and rooms from OTHER classes to prevent cross-class conflicts
        cross_class_entries = Timetable.query.filter(Timetable.class_id != self.class_id).all()
        for entry in cross_class_entries:
            if entry.faculty_id:
                faculty_busy[(entry.faculty_id, entry.day, entry.slot_number)] = True
            if entry.room_id:
                room_busy[(entry.room_id, entry.day, entry.slot_number)] = True

        # Per-subject per-day lecture count to distribute across days
        subject_day  = {}   # (subject_id, day) → count

        # ------------------------------------------------------------------
        # STEP 1 — BREAKS
        # ------------------------------------------------------------------
        for day in days:
            for s, (start, end) in enumerate(slots):
                if s in self.BREAK_SLOTS:
                    db.session.add(Timetable(
                        class_id=self.class_id,
                        day=day,
                        slot_number=s,
                        start_time=start,
                        end_time=end,
                        session_type='Break',
                        is_break=True
                    ))
                    class_busy[(day, s)] = True
        db.session.commit()

        # ------------------------------------------------------------------
        # STEP 2 — PRACTICALS  (parallel batches, same day+slot block)
        # ------------------------------------------------------------------
        # Build list of (class_subject) per practical session needed
        # Each cs appears practical_slots_per_week times
        practical_subjects = []   # only subjects with practicals
        for cs in self.class_subjects:
            if cs.practical_slots_per_week and cs.practical_slots_per_week > 0:
                for _ in range(cs.practical_slots_per_week):
                    practical_subjects.append(cs)

        # We need at least n_batches subjects per placement round
        # Each round consumes n_batches entries (one per batch)
        # Build rounds: group practical_subjects into sets of n_batches
        # Each set must have all-different subject_ids AND all-different faculty_ids
        # Strategy: shuffle and verify uniqueness; retry up to MAX_ROUND_TRIES
        MAX_ROUND_TRIES = 200

        def make_rounds(pool, n):
            """Split pool into groups of n with unique subject & faculty per group."""
            rounds = []
            remaining = list(pool)
            while len(remaining) >= n:
                placed = False
                for _ in range(MAX_ROUND_TRIES):
                    random.shuffle(remaining)
                    group = remaining[:n]
                    subj_ids = [g.subject_id for g in group]
                    fac_ids  = [g.faculty_id  for g in group]
                    if len(set(subj_ids)) == n and len(set(fac_ids)) == n:
                        rounds.append(group)
                        remaining = remaining[n:]
                        placed = True
                        break
                if not placed:
                    print(f"⚠ Could not form valid practical group from remaining {len(remaining)} subjects")
                    break
            return rounds

        practical_rounds = make_rounds(practical_subjects, n_batches)

        if len(self.labs) < n_batches:
            raise Exception(f"Need at least {n_batches} labs for parallel practicals, found {len(self.labs)}")

        for group in practical_rounds:
            assigned = False
            day_order = random.sample(days, len(days))
            for day in day_order:
                # Try all valid consecutive pairs (morning + afternoon both allowed)
                pair_order = list(self.practical_pairs)
                # Prefer afternoon pairs (slot index >= 4) by sorting descending
                pair_order.sort(key=lambda p: (-p[0]))  # higher index first (afternoon preference)
                for (s1, s2) in pair_order:
                    # Both slots in pair must be free for the class
                    if (day, s1) in class_busy or (day, s2) in class_busy:
                        continue

                    # Check all batches' faculty are free on both slots
                    ok = True
                    for cs in group:
                        for s in (s1, s2):
                            if (cs.faculty_id, day, s) in faculty_busy:
                                ok = False
                                break
                        if not ok:
                            break
                    if not ok:
                        continue

                    # Assign labs (one unique lab per batch)
                    labs_copy = list(self.labs)
                    random.shuffle(labs_copy)
                    lab_assigned = labs_copy[:n_batches]

                    # Check labs are free
                    lab_ok = True
                    for lab in lab_assigned:
                        for s in (s1, s2):
                            if (lab.id, day, s) in room_busy:
                                lab_ok = False
                                break
                        if not lab_ok:
                            break
                    if not lab_ok:
                        continue

                    # All checks passed — commit this practical round
                    for idx, cs in enumerate(group):
                        batch = self.batches[idx]
                        lab   = lab_assigned[idx]
                        for s in (s1, s2):
                            db.session.add(Timetable(
                                class_id=self.class_id,
                                day=day,
                                slot_number=s,
                                start_time=slots[s][0],
                                end_time=slots[s][1],
                                subject_id=cs.subject_id,
                                faculty_id=cs.faculty_id,
                                room_id=lab.id,
                                batch_id=batch.id,
                                session_type='Practical',
                                is_break=False
                            ))
                            faculty_busy[(cs.faculty_id, day, s)] = True
                            room_busy[(lab.id, day, s)] = True
                        # Mark class busy for both slots
                        class_busy[(day, s1)] = True
                        class_busy[(day, s2)] = True

                    assigned = True
                    break
                if assigned:
                    break

            if not assigned:
                print(f"⚠ Could not place practical round: {[cs.subject_id for cs in group]}")

        db.session.commit()

        # ------------------------------------------------------------------
        # STEP 3 — MENTORING  (one slot for the whole class, no subject)
        # ------------------------------------------------------------------
        mentoring_placed = False
        for day in days:
            if mentoring_placed:
                break
            for s, (start, end) in enumerate(slots):
                if s in self.BREAK_SLOTS:
                    continue
                if (day, s) in class_busy:
                    continue
                db.session.add(Timetable(
                    class_id=self.class_id,
                    day=day,
                    slot_number=s,
                    start_time=start,
                    end_time=end,
                    session_type='Mentoring',
                    is_break=False
                    # subject_id and faculty_id intentionally omitted
                ))
                class_busy[(day, s)] = True
                mentoring_placed = True
                break
        db.session.commit()

        # ------------------------------------------------------------------
        # STEP 4 — LECTURES  (distributed across days, no clustering)
        # ------------------------------------------------------------------
        # Build lecture pool: each cs repeated lecture_slots_per_week times
        lecture_pool = []
        for cs in self.class_subjects:
            for _ in range(cs.lecture_slots_per_week):
                lecture_pool.append(cs)
        random.shuffle(lecture_pool)

        for cs in lecture_pool:
            assigned = False
            # Randomise day order; prefer days where this subject hasn't been placed yet
            def day_sort_key(d):
                return subject_day.get((cs.subject_id, d), 0)

            day_order = sorted(random.sample(days, len(days)), key=day_sort_key)

            for day in day_order:
                if assigned:
                    break
                
                # Randomize slot order to avoid clustering all lectures in early slots
                slot_order = list(self.valid_slots)
                random.shuffle(slot_order)
                
                for s in slot_order:
                    start, end = slots[s]
                    if (day, s) in class_busy:
                        continue
                    if cs.faculty_id and (cs.faculty_id, day, s) in faculty_busy:
                        continue

                    # Find a free classroom
                    classroom = None
                    for room in self.classrooms:
                        if (room.id, day, s) not in room_busy:
                            classroom = room
                            break
                    if not classroom:
                        continue

                    db.session.add(Timetable(
                        class_id=self.class_id,
                        day=day,
                        slot_number=s,
                        start_time=start,
                        end_time=end,
                        subject_id=cs.subject_id,
                        faculty_id=cs.faculty_id,
                        room_id=classroom.id,
                        session_type='Lecture',
                        is_break=False
                    ))

                    if cs.faculty_id:
                        faculty_busy[(cs.faculty_id, day, s)] = True
                    room_busy[(classroom.id, day, s)] = True
                    class_busy[(day, s)] = True
                    subject_day[(cs.subject_id, day)] = subject_day.get((cs.subject_id, day), 0) + 1

                    assigned = True
                    break

            if not assigned:
                print(f"⚠ Could not place lecture for subject {cs.subject_id}")

        db.session.commit()
        print(f"✅ Timetable generated for class {self.class_id}")

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            flash('Logged in successfully!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if password != confirm_password:
            flash('Passwords do not match', 'error')
            return redirect(url_for('register'))
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'error')
            return redirect(url_for('register'))
        
        if User.query.filter_by(email=email).first():
            flash('Email already exists', 'error')
            return redirect(url_for('register'))
        
        user = User(username=username, email=email, is_admin=True)
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! Please login.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully', 'success')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    # Get statistics
    dept_count = Department.query.count()
    class_count = Class.query.count()
    faculty_count = Faculty.query.count()
    subject_count = Subject.query.count()
    
    # Get recent activity
    recent_classes = Class.query.order_by(Class.created_at.desc()).limit(5).all()
    
    return render_template('dashboard.html',
                         dept_count=dept_count,
                         class_count=class_count,
                         faculty_count=faculty_count,
                         subject_count=subject_count,
                         recent_classes=recent_classes)

# Department Management
@app.route('/departments')
@login_required
def manage_departments():
    departments = Department.query.all()
    return render_template('departments.html', departments=departments)

@app.route('/departments/add', methods=['POST'])
@login_required
def add_department():
    code = request.form.get('code')
    name = request.form.get('name')
    description = request.form.get('description')
    
    if Department.query.filter_by(code=code).first():
        flash('Department code already exists', 'error')
        return redirect(url_for('manage_departments'))
    
    department = Department(code=code, name=name, description=description)
    db.session.add(department)
    db.session.commit()
    
    flash('Department added successfully', 'success')
    return redirect(url_for('manage_departments'))

@app.route('/departments/edit/<int:id>', methods=['POST'])
@login_required
def edit_department(id):
    department = Department.query.get_or_404(id)
    code = request.form.get('code')
    name = request.form.get('name')
    description = request.form.get('description')
    
    existing = Department.query.filter_by(code=code).first()
    if existing and existing.id != id:
        flash('Department code already exists', 'error')
        return redirect(url_for('manage_departments'))
        
    department.code = code
    department.name = name
    department.description = description
    db.session.commit()
    flash('Department updated successfully', 'success')
    return redirect(url_for('manage_departments'))

@app.route('/departments/delete/<int:id>', methods=['POST'])
@login_required
def delete_department(id):
    department = Department.query.get_or_404(id)
    if department.classes or department.faculty or Subject.query.filter_by(department_id=id).first() or Room.query.filter_by(department_id=id).first():
        flash('Cannot delete department: It is assigned to classes, faculty, subjects, or rooms.', 'error')
        return redirect(url_for('manage_departments'))
    db.session.delete(department)
    db.session.commit()
    flash('Department deleted successfully', 'success')
    return redirect(url_for('manage_departments'))

# Class Management
@app.route('/classes')
@login_required
def manage_classes():
    classes = Class.query.all()
    departments = Department.query.all()
    semesters = list(range(1, 7))
    return render_template('classes.html', classes=classes, departments=departments, semesters=semesters)

@app.route('/classes/add', methods=['POST'])
@login_required
def add_class():
    name = request.form.get('name')
    code = request.form.get('code')
    year = request.form.get('year')
    department_id = request.form.get('department_id')
    semester = request.form.get('semester')
    strength = request.form.get('strength', 60)
    
    if Class.query.filter_by(code=code).first():
        flash('Class code already exists', 'error')
        return redirect(url_for('manage_classes'))
    
    class_obj = Class(
        name=name,
        code=code,
        year=year,
        department_id=department_id,
        semester=semester,
        strength=strength
    )
    
    db.session.add(class_obj)
    db.session.flush()  # Get the ID
    
    # Create batches (TB1, TB2, TB3)
    for i in range(1, 4):
        batch = Batch(
            name=f'TB{i}',
            code=f'{code}_TB{i}',
            class_id=class_obj.id
        )
        db.session.add(batch)
    
    db.session.commit()
    
    flash('Class and batches created successfully', 'success')
    return redirect(url_for('manage_classes'))

@app.route('/classes/edit/<int:id>', methods=['POST'])
@login_required
def edit_class(id):
    class_obj = Class.query.get_or_404(id)
    name = request.form.get('name')
    code = request.form.get('code')
    year = request.form.get('year')
    department_id = request.form.get('department_id')
    semester = request.form.get('semester')
    strength = request.form.get('strength', 60)
    
    existing = Class.query.filter_by(code=code).first()
    if existing and existing.id != id:
        flash('Class code already exists', 'error')
        return redirect(url_for('manage_classes'))
        
    class_obj.name = name
    class_obj.code = code
    class_obj.year = year
    class_obj.department_id = department_id
    class_obj.semester = semester
    class_obj.strength = strength
    db.session.commit()
    flash('Class updated successfully', 'success')
    return redirect(url_for('manage_classes'))

@app.route('/classes/delete/<int:id>', methods=['POST'])
@login_required
def delete_class(id):
    class_obj = Class.query.get_or_404(id)
    if class_obj.timetable:
        flash('Cannot delete class: It has an active timetable generated.', 'error')
        return redirect(url_for('manage_classes'))
    
    Batch.query.filter_by(class_id=id).delete()
    ClassSubject.query.filter_by(class_id=id).delete()
    PracticalSlot.query.filter_by(class_id=id).delete()
    db.session.delete(class_obj)
    db.session.commit()
    flash('Class and its batches deleted successfully', 'success')
    return redirect(url_for('manage_classes'))


# Subject Management
@app.route('/subjects')
@login_required
def manage_subjects():
    subjects = Subject.query.all()
    departments = Department.query.all()
    return render_template('subjects.html', subjects=subjects, departments=departments)

@app.route('/subjects/add', methods=['POST'])
@login_required
def add_subject():
    code = request.form.get('code')
    name = request.form.get('name')
    type = request.form.get('type')
    lecture_hours = request.form.get('lecture_hours', 0, type=int)
    practical_hours = request.form.get('practical_hours', 0, type=int)
    credits = request.form.get('credits', 3, type=int)
    department_id = request.form.get('department_id')
    
    if Subject.query.filter_by(code=code).first():
        flash('Subject code already exists', 'error')
        return redirect(url_for('manage_subjects'))
    
    subject = Subject(
        code=code,
        name=name,
        type=type,
        lecture_hours=lecture_hours,
        practical_hours=practical_hours,
        credits=credits,
        department_id=department_id if department_id else None
    )
    
    db.session.add(subject)
    db.session.commit()
    
    flash('Subject added successfully', 'success')
    return redirect(url_for('manage_subjects'))

@app.route('/subjects/edit/<int:id>', methods=['POST'])
@login_required
def edit_subject(id):
    subject = Subject.query.get_or_404(id)
    code = request.form.get('code')
    name = request.form.get('name')
    type = request.form.get('type')
    lecture_hours = request.form.get('lecture_hours', 0, type=int)
    practical_hours = request.form.get('practical_hours', 0, type=int)
    credits = request.form.get('credits', 3, type=int)
    department_id = request.form.get('department_id')
    
    existing = Subject.query.filter_by(code=code).first()
    if existing and existing.id != id:
        flash('Subject code already exists', 'error')
        return redirect(url_for('manage_subjects'))
        
    subject.code = code
    subject.name = name
    subject.type = type
    subject.lecture_hours = lecture_hours
    subject.practical_hours = practical_hours
    subject.credits = credits
    subject.department_id = department_id if department_id else None
    db.session.commit()
    flash('Subject updated successfully', 'success')
    return redirect(url_for('manage_subjects'))

@app.route('/subjects/delete/<int:id>', methods=['POST'])
@login_required
def delete_subject(id):
    subject = Subject.query.get_or_404(id)
    if subject.class_subjects or Timetable.query.filter_by(subject_id=id).first():
        flash('Cannot delete subject: It is assigned to classes or a timetable.', 'error')
        return redirect(url_for('manage_subjects'))
    db.session.delete(subject)
    db.session.commit()
    flash('Subject deleted successfully', 'success')
    return redirect(url_for('manage_subjects'))


# Faculty Management
@app.route('/faculty')
@login_required
def manage_faculty():
    faculty_list = Faculty.query.all()
    departments = Department.query.all()
    return render_template('faculty.html', faculty_list=faculty_list, departments=departments)

@app.route('/faculty/add', methods=['POST'])
@login_required
def add_faculty():
    employee_id = request.form.get('employee_id')
    name = request.form.get('name')
    email = request.form.get('email')
    phone = request.form.get('phone')
    department_id = request.form.get('department_id')
    designation = request.form.get('designation')
    qualification = request.form.get('qualification')
    
    if Faculty.query.filter_by(employee_id=employee_id).first():
        flash('Employee ID already exists', 'error')
        return redirect(url_for('manage_faculty'))
    
    if Faculty.query.filter_by(email=email).first():
        flash('Email already exists', 'error')
        return redirect(url_for('manage_faculty'))
    
    faculty = Faculty(
        employee_id=employee_id,
        name=name,
        email=email,
        phone=phone,
        department_id=department_id if department_id else None,
        designation=designation,
        qualification=qualification
    )
    
    db.session.add(faculty)
    db.session.commit()
    
    flash('Faculty added successfully', 'success')
    return redirect(url_for('manage_faculty'))

@app.route('/faculty/edit/<int:id>', methods=['POST'])
@login_required
def edit_faculty(id):
    faculty = Faculty.query.get_or_404(id)
    employee_id = request.form.get('employee_id')
    name = request.form.get('name')
    email = request.form.get('email')
    phone = request.form.get('phone')
    department_id = request.form.get('department_id')
    designation = request.form.get('designation')
    qualification = request.form.get('qualification')
    
    existing_emp = Faculty.query.filter_by(employee_id=employee_id).first()
    if existing_emp and existing_emp.id != id:
        flash('Employee ID already exists', 'error')
        return redirect(url_for('manage_faculty'))
        
    existing_email = Faculty.query.filter_by(email=email).first()
    if existing_email and existing_email.id != id:
        flash('Email already exists', 'error')
        return redirect(url_for('manage_faculty'))
        
    faculty.employee_id = employee_id
    faculty.name = name
    faculty.email = email
    faculty.phone = phone
    faculty.department_id = department_id if department_id else None
    faculty.designation = designation
    faculty.qualification = qualification
    db.session.commit()
    flash('Faculty updated successfully', 'success')
    return redirect(url_for('manage_faculty'))

@app.route('/faculty/delete/<int:id>', methods=['POST'])
@login_required
def delete_faculty(id):
    faculty = Faculty.query.get_or_404(id)
    if faculty.class_subjects or faculty.mentored_batches or faculty.timetable_entries:
        flash('Cannot delete faculty: They are assigned to subjects, batches, or a timetable.', 'error')
        return redirect(url_for('manage_faculty'))
    db.session.delete(faculty)
    db.session.commit()
    flash('Faculty deleted successfully', 'success')
    return redirect(url_for('manage_faculty'))


# Room Management
@app.route('/rooms')
@login_required
def manage_rooms():
    rooms = Room.query.all()
    departments = Department.query.all()
    return render_template('rooms.html', rooms=rooms, departments=departments)

@app.route('/rooms/add', methods=['POST'])
@login_required
def add_room():
    room_number = request.form.get('room_number')
    room_type = request.form.get('room_type')
    capacity = request.form.get('capacity', type=int)
    department_id = request.form.get('department_id')
    equipment = request.form.get('equipment')
    
    if Room.query.filter_by(room_number=room_number).first():
        flash('Room number already exists', 'error')
        return redirect(url_for('manage_rooms'))
    
    room = Room(
        room_number=room_number,
        room_type=room_type,
        capacity=capacity,
        department_id=department_id if department_id else None,
        equipment=equipment
    )
    
    db.session.add(room)
    db.session.commit()
    
    flash('Room added successfully', 'success')
    return redirect(url_for('manage_rooms'))

@app.route('/rooms/edit/<int:id>', methods=['POST'])
@login_required
def edit_room(id):
    room = Room.query.get_or_404(id)
    room_number = request.form.get('room_number')
    room_type = request.form.get('room_type')
    capacity = request.form.get('capacity', type=int)
    department_id = request.form.get('department_id')
    equipment = request.form.get('equipment')
    
    existing = Room.query.filter_by(room_number=room_number).first()
    if existing and existing.id != id:
        flash('Room number already exists', 'error')
        return redirect(url_for('manage_rooms'))
        
    room.room_number = room_number
    room.room_type = room_type
    room.capacity = capacity
    room.department_id = department_id if department_id else None
    room.equipment = equipment
    db.session.commit()
    flash('Room updated successfully', 'success')
    return redirect(url_for('manage_rooms'))

@app.route('/rooms/delete/<int:id>', methods=['POST'])
@login_required
def delete_room(id):
    room = Room.query.get_or_404(id)
    if room.timetable_entries:
        flash('Cannot delete room: It is used in an active timetable.', 'error')
        return redirect(url_for('manage_rooms'))
    db.session.delete(room)
    db.session.commit()
    flash('Room deleted successfully', 'success')
    return redirect(url_for('manage_rooms'))


# Class Subjects Assignment
@app.route('/class-subjects/<int:class_id>')
@login_required
def manage_class_subjects(class_id):
    class_obj = Class.query.get_or_404(class_id)
    subjects = Subject.query.all()
    faculty_list = Faculty.query.all()
    class_subjects = ClassSubject.query.filter_by(class_id=class_id).all()
    
    return render_template('class_subjects.html',
                         class_obj=class_obj,
                         subjects=subjects,
                         faculty_list=faculty_list,
                         class_subjects=class_subjects)

@app.route('/class-subjects/add', methods=['POST'])
@login_required
def add_class_subject():
    class_id = request.form.get('class_id')
    subject_id = request.form.get('subject_id')
    faculty_id = request.form.get('faculty_id')
    lecture_slots = request.form.get('lecture_slots', 3, type=int)
    practical_slots = request.form.get('practical_slots', 2, type=int)
    
    # Check if already assigned
    existing = ClassSubject.query.filter_by(class_id=class_id, subject_id=subject_id).first()
    if existing:
        flash('Subject already assigned to this class', 'error')
        return redirect(url_for('manage_class_subjects', class_id=class_id))
    
    class_subject = ClassSubject(
        class_id=class_id,
        subject_id=subject_id,
        faculty_id=faculty_id,
        lecture_slots_per_week=lecture_slots,
        practical_slots_per_week=practical_slots
    )
    
    db.session.add(class_subject)
    db.session.commit()
    
    flash('Subject assigned successfully', 'success')
    return redirect(url_for('manage_class_subjects', class_id=class_id))

@app.route('/class-subjects/update/<int:id>', methods=['POST'])
@login_required
def update_class_subject(id):

    class_subject = ClassSubject.query.get_or_404(id)

    class_subject.faculty_id = request.form.get('faculty_id')
    class_subject.lecture_slots_per_week = request.form.get('lecture_slots', type=int)
    class_subject.practical_slots_per_week = request.form.get('practical_slots', type=int)

    db.session.commit()

    flash('Subject assignment updated successfully', 'success')
    return redirect(url_for('manage_class_subjects', class_id=class_subject.class_id))



@app.route('/class-subjects/delete/<int:id>')
@login_required
def delete_class_subject(id):

    class_subject = ClassSubject.query.get_or_404(id)
    class_id = class_subject.class_id

    db.session.delete(class_subject)
    db.session.commit()

    flash('Subject assignment deleted successfully', 'success')
    return redirect(url_for('manage_class_subjects', class_id=class_id))


# Batch Mentors Assignment
@app.route('/batch-mentors/<int:class_id>')
@login_required
def manage_batch_mentors(class_id):
    class_obj = Class.query.get_or_404(class_id)
    batches = Batch.query.filter_by(class_id=class_id).all()
    faculty_list = Faculty.query.all()
    
    return render_template('batch_mentors.html',
                         class_obj=class_obj,
                         batches=batches,
                         faculty_list=faculty_list)

@app.route('/batch-mentors/assign', methods=['POST'])
@login_required
def assign_batch_mentor():
    batch_id = request.form.get('batch_id')
    mentor_id = request.form.get('mentor_id')
    
    batch = Batch.query.get_or_404(batch_id)
    batch.mentor_id = mentor_id if mentor_id else None
    
    db.session.commit()
    
    flash('Mentor assigned successfully', 'success')
    return redirect(url_for('manage_batch_mentors', class_id=batch.class_id))

# Generate Timetable
@app.route('/generate-timetable')
@login_required
def generate_timetable():
    classes = Class.query.all()
    return render_template('generate.html', classes=classes)

@app.route('/generate-timetable/run/<int:class_id>')
@login_required
def run_timetable_generation(class_id):
    try:
        # Check if class exists
        class_obj = Class.query.get_or_404(class_id)
        
        # Check if class has subjects assigned
        if not ClassSubject.query.filter_by(class_id=class_id).first():
            flash('Please assign subjects to this class first', 'error')
            return redirect(url_for('manage_class_subjects', class_id=class_id))
        
        # Check if rooms exist
        rooms = Room.query.all()
        if not rooms:
            flash('Please add classrooms and labs first', 'error')
            return redirect(url_for('manage_rooms'))
        
        # Initialize and run genetic algorithm
        ga = GeneticAlgorithmTimetable(class_id)
        ga.generate_structured_timetable()
        flash('Structured timetable generated successfully!', 'success')

    except Exception as e:
        flash(f'Error generating timetable: {str(e)}', 'error')
        import traceback
        print(traceback.format_exc())
    
    return redirect(url_for('view_timetable') + f'?class_id={class_id}')

# View Timetable
@app.route('/view-timetable')
@login_required
def view_timetable():
    class_id = request.args.get('class_id', type=int)
    classes = Class.query.all()
    
    timetable_data = None
    time_slots = app.config['TIME_SLOTS']
    
    if class_id:
        # Get timetable for selected class
        timetable_entries = Timetable.query.filter_by(class_id=class_id)\
            .order_by(Timetable.day, Timetable.slot_number).all()
        print("Timetable entries found:", len(timetable_entries))
        for e in timetable_entries:
            print("DAY:", e.day, "SLOT:", e.slot_number, "TYPE:", e.session_type)
        # Organize by day and slot
        timetable_data = {}
        days = app.config['DAYS']
        entries = Timetable.query.filter_by(class_id=class_id).all()
        if entries:
            # Initialize with empty lists
            timetable_data = {day: [[] for _ in range(len(app.config['TIME_SLOTS']))] for day in app.config['DAYS']}
            for entry in entries:
                timetable_data[entry.day][entry.slot_number].append(entry)
                
    return render_template('timetable.html', 
                         classes=classes, 
                         class_id=class_id,
                         timetable_data=timetable_data,
                         time_slots=app.config['TIME_SLOTS'],
                         days=app.config['DAYS'])

@app.route('/export_timetable/<int:class_id>')
@login_required
def export_timetable(class_id):
    cls = Class.query.get_or_404(class_id)
    timetable_entries = Timetable.query.filter_by(class_id=class_id).all()
    
    if not timetable_entries:
        flash('No timetable generated for this class yet.', 'error')
        return redirect(url_for('view_timetable', class_id=class_id))
    
    # Process timetable data exactly like view_timetable does
    timetable_data = {day: [[] for _ in range(len(app.config['TIME_SLOTS']))] for day in app.config['DAYS']}
    for entry in timetable_entries:
        timetable_data[entry.day][entry.slot_number].append(entry)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    # Styles
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    heavy_side = Side(border_style="medium", color="000000")
    heavy_border = Border(left=heavy_side, right=heavy_side, top=heavy_side, bottom=heavy_side)
    gray_fill = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")
    
    # Set Column Widths
    ws.column_dimensions['A'].width = 20  # Time
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 25
        
    current_row = 1
    
    # Headers
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws.cell(row=current_row, column=1)
    cell.value = "DKTE SOCIETY'S"
    cell.font = bold_font
    cell.alignment = center_align
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws.cell(row=current_row, column=1)
    cell.value = "YASHWANTRAO CHAVAN POLYTECHNIC, ICHALKARANJI"
    cell.font = title_font
    cell.alignment = center_align
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws.cell(row=current_row, column=1)
    cell.value = f"DEPARTMENT OF {cls.department.name.upper()}"
    cell.font = subtitle_font
    cell.alignment = center_align
    current_row += 2
    
    # Subheader row
    ws.merge_cells(f'D{current_row}:E{current_row}')
    cell = ws.cell(row=current_row, column=4)
    cell.value = "TIME TABLE 2025-26"
    cell.font = bold_font
    cell.alignment = center_align
    
    ws.merge_cells(f'F{current_row}:G{current_row}')
    cell = ws.cell(row=current_row, column=6)
    cell.value = f"CLASS: {cls.code}"
    cell.font = bold_font
    cell.alignment = center_align
    current_row += 1
    
    # Table Headers
    headers = ["TIME", "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header_text
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = heavy_border
    current_row += 1
    
    # Table rows
    time_slots = app.config['TIME_SLOTS']
    for slot_idx in range(len(time_slots)):
        is_break = False
        try:
            if timetable_data['Monday'][slot_idx] and timetable_data['Monday'][slot_idx][0].is_break:
                is_break = True
        except:
            pass
            
        # Write Time Column
        time_str = f"{time_slots[slot_idx][0]} TO {time_slots[slot_idx][1]}"
        cell = ws.cell(row=current_row, column=1)
        cell.value = time_str
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        
        if is_break:
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
            break_cell = ws.cell(row=current_row, column=2)
            break_cell.value = "LUNCH BREAK" if slot_idx == 3 else "SHORT BREAK"
            break_cell.font = bold_font
            break_cell.alignment = center_align
            break_cell.fill = gray_fill
            for c in range(2, 7):
                ws.cell(row=current_row, column=c).border = thin_border
            
            sat_cell = ws.cell(row=current_row, column=7)
            sat_cell.border = thin_border
            sat_cell.fill = gray_fill
        else:
            for day_idx, day in enumerate(app.config['DAYS'], 2):
                cell = ws.cell(row=current_row, column=day_idx)
                cell.border = thin_border
                cell.alignment = center_align
                entries = timetable_data[day][slot_idx]
                if entries:
                    if entries[0].session_type == 'Mentoring':
                        cell.value = "MENTORING"
                        cell.font = bold_font
                    elif entries[0].session_type == 'Lecture':
                        subject_code = entries[0].subject.code
                        faculty_name = entries[0].faculty_ref.name if entries[0].faculty_ref else ""
                        room_number = entries[0].room.room_number if entries[0].room_id else ""
                        cell.value = f"{subject_code}\n({faculty_name})\n({room_number})"
                    elif entries[0].session_type == 'Practical':
                        parts = []
                        for entry in entries:
                            fac_code = entry.room.room_number if entry.room_id else ""
                            parts.append(f"{entry.subject.code}-{entry.batch.name} ({fac_code})")
                        cell.value = "\n".join(parts)
            
            sat_cell = ws.cell(row=current_row, column=7)
            sat_cell.border = thin_border
        
        ws.row_dimensions[current_row].height = 40
        current_row += 1
        
    current_row += 2
    subject_headers = ["SUBJECT", "FACULTY", "SUBJECT", "FACULTY"]
    for col_idx, header_text in enumerate(subject_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header_text
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
    current_row += 1
    
    unique_subjects = list({cs.subject_id: cs for cs in cls.subjects}.values())
    batches_of_two = [unique_subjects[i:i + 2] for i in range(0, len(unique_subjects), 2)]
    
    for row in batches_of_two:
        c1 = ws.cell(row=current_row, column=1)
        c1.value = row[0].subject_ref.name if len(row) > 0 else ""
        c1.alignment = left_align
        c1.border = thin_border
        
        c2 = ws.cell(row=current_row, column=2)
        c2.value = row[0].faculty_ref.name if len(row) > 0 and row[0].faculty_ref else ""
        c2.alignment = left_align
        c2.border = thin_border
        
        c3 = ws.cell(row=current_row, column=3)
        c3.value = row[1].subject_ref.name if len(row) > 1 else ""
        c3.alignment = left_align
        c3.border = thin_border
        
        c4 = ws.cell(row=current_row, column=4)
        c4.value = row[1].faculty_ref.name if len(row) > 1 and row[1].faculty_ref else ""
        c4.alignment = left_align
        c4.border = thin_border
        
        current_row += 1
        
    current_row += 3
    sigs = ["Time Table Incharge", "HOD", "Vice-Principal", "Principal"]
    for i, sig in enumerate(sigs):
        col = (i * 2) + 1
        if col > 7: col = 7
        cell = ws.cell(row=current_row, column=col)
        cell.value = sig
        cell.font = bold_font
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"Timetable_{cls.code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(out, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Initialize Sample Data
@app.route('/init-sample-data')
def init_sample_data():
    """Initialize with sample data for testing"""
    # Clear existing data
    db.session.query(Timetable).delete()
    db.session.query(PracticalSlot).delete()
    db.session.query(ClassSubject).delete()
    db.session.query(Batch).delete()
    db.session.query(Class).delete()
    db.session.query(Faculty).delete()
    db.session.query(Subject).delete()
    db.session.query(Room).delete()
    db.session.query(Department).delete()

    # ------------------------------------------------------------------
    # Departments
    # ------------------------------------------------------------------
    dept1 = Department(code='CSE', name='Computer Engineering',
                       description='Computer Science and Engineering Department')
    dept2 = Department(code='MECH', name='Mechanical Engineering',
                       description='Mechanical Engineering Department')
    db.session.add_all([dept1, dept2])
    db.session.commit()

    # ------------------------------------------------------------------
    # Faculty  (3 unique faculty for CSE — one per batch practical)
    # ------------------------------------------------------------------
    faculty1 = Faculty(employee_id='F001', name='S.S. Mali',
                       email='mali@college.edu', phone='9876543210',
                       department_id=dept1.id, designation='Professor',
                       qualification='Ph.D. Computer Science')
    faculty2 = Faculty(employee_id='F002', name='P.N. Patil',
                       email='patil@college.edu', phone='9876543211',
                       department_id=dept1.id, designation='Associate Professor',
                       qualification='M.Tech CSE')
    faculty3 = Faculty(employee_id='F003', name='S.N. Kumbhar',
                       email='kumbhar@college.edu', phone='9876543212',
                       department_id=dept1.id, designation='Assistant Professor',
                       qualification='M.Tech CSE')
    db.session.add_all([faculty1, faculty2, faculty3])
    db.session.commit()

    # ------------------------------------------------------------------
    # Rooms — 2 classrooms + 3 labs (one lab per batch for parallel practicals)
    # ------------------------------------------------------------------
    room_cr1 = Room(room_number='A-101', room_type='Classroom', capacity=60, department_id=dept1.id)
    room_cr2 = Room(room_number='A-102', room_type='Classroom', capacity=60, department_id=dept1.id)
    room_lab1 = Room(room_number='LAB-1', room_type='Lab', capacity=30, department_id=dept1.id,
                     equipment='30 Computers, Projector')
    room_lab2 = Room(room_number='LAB-2', room_type='Lab', capacity=30, department_id=dept1.id,
                     equipment='30 Computers, Projector')
    room_lab3 = Room(room_number='LAB-3', room_type='Lab', capacity=30, department_id=dept1.id,
                     equipment='30 Computers, Projector')
    db.session.add_all([room_cr1, room_cr2, room_lab1, room_lab2, room_lab3])
    db.session.commit()

    # ------------------------------------------------------------------
    # Subjects — 3 with practicals (one per batch), 1 theory-only
    # ------------------------------------------------------------------
    sub_os  = Subject(code='OSY',  name='Operating System',            type='Theory',
                      lecture_hours=3, practical_hours=2, credits=4, department_id=dept1.id)
    sub_ste = Subject(code='STE',  name='Software Testing',            type='Theory',
                      lecture_hours=3, practical_hours=2, credits=4, department_id=dept1.id)
    sub_dan = Subject(code='DAN',  name='Data Analytics',              type='Theory',
                      lecture_hours=3, practical_hours=2, credits=4, department_id=dept1.id)
    sub_ede = Subject(code='EDE',  name='Entrepreneurship Development', type='Theory',
                      lecture_hours=3, practical_hours=0, credits=3, department_id=dept1.id)
    db.session.add_all([sub_os, sub_ste, sub_dan, sub_ede])
    db.session.commit()

    # ------------------------------------------------------------------
    # Class  TYCW-A (matching the reference timetable image)
    # ------------------------------------------------------------------
    class1 = Class(name='Third Year Computer Engineering A',
                   code='TYCW-A', year='TY',
                   department_id=dept1.id, semester=5, strength=60)
    db.session.add(class1)
    db.session.flush()

    # Create 3 batches TB1, TB2, TB3
    batch1 = Batch(name='TB1', code='TYCW-A_TB1', class_id=class1.id, mentor_id=faculty1.id)
    batch2 = Batch(name='TB2', code='TYCW-A_TB2', class_id=class1.id, mentor_id=faculty2.id)
    batch3 = Batch(name='TB3', code='TYCW-A_TB3', class_id=class1.id, mentor_id=faculty3.id)
    db.session.add_all([batch1, batch2, batch3])
    db.session.commit()

    # ------------------------------------------------------------------
    # Assign subjects to class  (3 subjects with practicals, 1 without)
    # Each practical subject has a DIFFERENT faculty — required for parallel scheduling
    # ------------------------------------------------------------------
    cs_os  = ClassSubject(class_id=class1.id, subject_id=sub_os.id,  faculty_id=faculty1.id,
                          lecture_slots_per_week=3, practical_slots_per_week=1)
    cs_ste = ClassSubject(class_id=class1.id, subject_id=sub_ste.id, faculty_id=faculty2.id,
                          lecture_slots_per_week=3, practical_slots_per_week=1)
    cs_dan = ClassSubject(class_id=class1.id, subject_id=sub_dan.id, faculty_id=faculty3.id,
                          lecture_slots_per_week=3, practical_slots_per_week=1)
    cs_ede = ClassSubject(class_id=class1.id, subject_id=sub_ede.id, faculty_id=faculty1.id,
                          lecture_slots_per_week=3, practical_slots_per_week=0)
    db.session.add_all([cs_os, cs_ste, cs_dan, cs_ede])
    db.session.commit()

    flash('Sample data initialized successfully! Class: TYCW-A | Subjects: OSY, STE, DAN, EDE', 'success')
    return redirect(url_for('dashboard'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)